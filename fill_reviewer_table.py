import json
import os
import openpyxl

XLS_FILE = 'Abstracts_with_reviewer_assignments.xlsx'
OUTPUT_FILE = os.path.splitext(XLS_FILE)[0] + '_filled.xlsx'
HEADER_ROWS = 1
ID_COLUMN = 1
REVIEWER_COLUMNS = [3, 4, 5]

workbook = openpyxl.load_workbook(XLS_FILE)
sheet = workbook.active

current_row = HEADER_ROWS + 1

with open('reviewer_assignments.json', 'r') as f:
    reviewer_assignments = json.load(f)

assignments_dict = {
    a['abstract_number']: a for a in reviewer_assignments
}

abstract_id = sheet.cell(row=current_row, column=ID_COLUMN).value
while abstract_id:
    print(f"Processing abstract ID: {abstract_id}")
    id_key = '#' + str(abstract_id)
    for reviewer_n, reviewer in enumerate(assignments_dict[id_key]['assigned_reviewers']):
        reviewer_str = f'{reviewer["reviewer_name"][1]}, {reviewer["reviewer_name"][0]}'
        sheet.cell(row=current_row, column=REVIEWER_COLUMNS[reviewer_n], value=reviewer_str)

    current_row += 1
    abstract_id = sheet.cell(row=current_row, column=ID_COLUMN).value


workbook.save(OUTPUT_FILE)